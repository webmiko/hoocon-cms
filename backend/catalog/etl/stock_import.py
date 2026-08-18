"""Import warehouse stock quantities from a 1C Excel export.

Expected columns (first sheet): ``Артикул`` and qty as ``Свободно`` /
``Остатки`` / ``Остаток``. Only catalog SKUs are updated; unknown артикулы
are ignored (merge, not wipe).
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, BinaryIO

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from catalog.models import SKU
from config.logging_utils import setup_logger

logger = setup_logger(__name__)

SKU_HEADER_CANON: frozenset[str] = frozenset({"артикул"})
QTY_HEADER_CANON: frozenset[str] = frozenset({"остаток", "остатки", "свободно"})

_HEADER_SPACE_RE = re.compile(r"\s+")
# 1C «BV232-A» / «BV232A» / CMS «8100-bv232a» → key BV232A.
_BV_BODY_RE = re.compile(
    r"(?i)(?:^|[^a-z0-9])(?:8100[-_]?)?bv[-_]?(\d{3,4})[-_]?([a-e])?(?=$|[^a-z0-9])",
)
# Factory kits H8101…H8122-BV… — not the bare valve body (series 81 cards).
_H81_KIT_RE = re.compile(r"(?i)^h81")
# H8205-LAV regulating valves — distinct from BV bodies / H81 kits.
_H8205_RE = re.compile(r"(?i)^h8205")
# Full electrical edition → body key for 1C bare articles (``H8121-BV265``).
_H81_EDITION_RE = re.compile(
    r"(?i)^(?P<body>h81(?:01|02|03|04|05|06|07|08|21|22)-bv\d{3,4}[a-e]?)"
    r"-(?:24|230)(?:as|a|ds|d)$",
)
_H81_BARE_RE = re.compile(
    r"(?i)^h81(?:01|02|03|04|05|06|07|08|21|22)-bv\d{3,4}[a-e]?$",
)
_H8205_EDITION_RE = re.compile(
    r"(?i)^(?P<body>h8205-lav[23](?:32|40|50|65|80|100|125|150|200|250|300)"
    r"(?:st|s|t)?)-(?P<tail>24|230)[adm]$",
)
_H8205_BARE_RE = re.compile(
    r"(?i)^h8205-lav[23](?:32|40|50|65|80|100|125|150|200|250|300)(?:st|s|t)?$",
)
_TRAILING_NOTE_RE = re.compile(
    r"(?i)\s+(?:DN\s*\d+|\d+\s*[.…\-−–—]+\s*\d+\s*[mм][aа].*)$",
)
# 1C warehouse line for the 4–20 mA factory option on the same sku_code.
# Latin/Cyrillic ``mA`` / ``мА`` is the discriminator (not ``HVA24-20``).
_MA_OPTION_RE = re.compile(
    r"(?i)(?:0\s*\(?\s*)?4\s*[-–—.]*\s*20\s*[mм][aа]\b",
)


@dataclass
class StockImportReport:
    """Counters from one stock XLSX apply."""

    updated: int = 0
    ignored_unknown: int = 0
    bad_rows: int = 0
    skipped_blank: int = 0
    unknown_codes: list[str] = field(default_factory=list)

    def summary(self) -> str:
        """Human-readable one-line report for Admin messages."""
        unknown_hint = ""
        if self.unknown_codes:
            sample = ", ".join(self.unknown_codes[:5])
            more = len(self.unknown_codes) - 5
            if more > 0:
                sample = f"{sample}… (+{more})"
            unknown_hint = f" (примеры: {sample})"
        return (
            f"Обновлено: {self.updated}; "
            f"неизвестных артикулов: {self.ignored_unknown}{unknown_hint}; "
            f"битых строк: {self.bad_rows}; "
            f"пустых: {self.skipped_blank}."
        )


class StockImportError(ValueError):
    """Invalid workbook structure (missing headers, empty sheet, …)."""


def normalize_stock_header(value: Any) -> str:
    """Collapse header cell to a comparable token (lowercase, no spaces)."""
    text = " ".join(str(value or "").strip().casefold().split())
    text = text.replace("ё", "е")
    return _HEADER_SPACE_RE.sub("", text)


def normalize_stock_article_key(raw: str) -> str:
    """Canonical match key for a stock / catalog article.

    Ball-valve bodies (series 81 cards): ``BV232-A``, ``BV232A``,
    ``8100-bv232a``, ``BV265 DN65`` → ``BV232A`` / ``BV265``.

    H81xx kits (``H8103-BV265-24A``, ``H8121-BV265``, …) keep a distinct key so
    their qty does not land on a bare ``BV*`` / ``8100-bv*`` body SKU.
    Bare 1C kit codes (``H8121-BV265``) fan out to all electrical editions
    in :func:`apply_stock_rows`.

    H8205-LAV editions keep a distinct key (never mapped onto BV bodies).
    Bare ``H8205-LAV280`` / ``H8205-LAV280ST`` fan out to matching editions.

    Other articles: casefold, strip ``8100-``, drop trailing notes
    (``4-20 mA``, ``DN65``). The 4–20 mA note is *not* a different SKU —
    see :func:`stock_article_is_ma_option`.

    Args:
        raw: Article from Excel or ``SKU.sku_code``.

    Returns:
        Uppercase key for dict lookup.
    """
    text = " ".join(str(raw or "").strip().split())
    if not text:
        return ""
    text = _TRAILING_NOTE_RE.sub("", text).strip()
    if _H81_KIT_RE.match(text) or _H8205_RE.match(text):
        return re.sub(r"[\s_]+", "-", text).upper()

    match = _BV_BODY_RE.search(f" {text} ")
    if match:
        letter = (match.group(2) or "").upper()
        return f"BV{match.group(1)}{letter}"

    key = text.casefold()
    if key.startswith("8100-"):
        key = key[5:]
    return re.sub(r"[\s_]+", "-", key).upper()


def stock_article_is_ma_option(raw: str) -> bool:
    """True when the 1C article is the 4–20 mA (special-order) warehouse line.

    The modulating SKU keeps one ``sku_code``; 1C adds a second row
    ``DA10FU24-AS 4-20 mA`` for units already switched to current mode.

    Args:
        raw: Article cell from the stock workbook.

    Returns:
        True when the cell names the 4–20 mA option, not the base article.
    """
    text = " ".join(str(raw or "").strip().split())
    if not text:
        return False
    return _MA_OPTION_RE.search(text) is not None


def h81_kit_bare_stock_key(normalized_key: str) -> str | None:
    """Return bare H81 kit key (``H8121-BV265``) for fan-out, or ``None``.

    Args:
        normalized_key: Output of :func:`normalize_stock_article_key`.

    Returns:
        Uppercase bare kit key when the article is an H81 body or edition;
        otherwise ``None``.
    """
    key = (normalized_key or "").strip().upper()
    if not key:
        return None
    edition = _H81_EDITION_RE.fullmatch(key)
    if edition is not None:
        return edition.group("body").upper()
    if _H81_BARE_RE.fullmatch(key):
        return key
    return None


def h8205_lav_bare_stock_key(normalized_key: str) -> str | None:
    """Return bare H8205 body key (``H8205-LAV280`` / ``…ST``) for fan-out.

    Args:
        normalized_key: Output of :func:`normalize_stock_article_key`.

    Returns:
        Uppercase bare LAV key when the article is an H8205 body or edition;
        otherwise ``None``.
    """
    key = (normalized_key or "").strip().upper()
    if not key:
        return None
    edition = _H8205_EDITION_RE.fullmatch(key)
    if edition is not None:
        return edition.group("body").upper()
    if _H8205_BARE_RE.fullmatch(key):
        return key
    return None


def parse_stock_qty(value: Any) -> int | None:
    """Parse a quantity cell; fractional values floor toward zero.

    Args:
        value: Cell value (int, float, or numeric string).

    Returns:
        Integer qty, or ``None`` when the cell is not a number.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return math.floor(value)
    text = str(value).strip().replace("\u00a0", " ").replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return math.floor(float(text))
    except ValueError:
        return None


def _find_header_indexes(header_row: tuple[Any, ...]) -> tuple[int, int]:
    """Return (sku_col, qty_col) 0-based indexes from the header row."""
    sku_idx: int | None = None
    qty_idx: int | None = None
    for idx, cell in enumerate(header_row):
        token = normalize_stock_header(cell)
        if not token:
            continue
        if token in SKU_HEADER_CANON and sku_idx is None:
            sku_idx = idx
        elif token in QTY_HEADER_CANON and qty_idx is None:
            qty_idx = idx
    if sku_idx is None or qty_idx is None:
        raise StockImportError(
            "В первой строке нужны заголовки «Артикул» и «Свободно» (или «Остатки» / «Остаток»).",
        )
    return sku_idx, qty_idx


def parse_stock_rows(file_obj: BinaryIO | BytesIO) -> list[tuple[str, int | None]]:
    """Read артикул + qty pairs from the first worksheet.

    Empty артикул rows are omitted. Non-numeric qty yields ``(code, None)``.

    Args:
        file_obj: Binary ``.xlsx`` stream (seekable).

    Returns:
        List of ``(sku_code, qty_or_None)``.

    Raises:
        StockImportError: missing required headers or unreadable workbook.
    """
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl raises many types
        raise StockImportError(f"Не удалось открыть Excel: {exc}") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise StockImportError("В книге нет активного листа.")
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise StockImportError("Файл пуст.") from exc
        sku_idx, qty_idx = _find_header_indexes(tuple(header or ()))

        parsed: list[tuple[str, int | None]] = []
        for row in rows_iter:
            if row is None:
                continue
            cells = tuple(row)
            sku_raw = cells[sku_idx] if sku_idx < len(cells) else None
            qty_raw = cells[qty_idx] if qty_idx < len(cells) else None
            code = str(sku_raw or "").strip()
            if not code:
                continue
            parsed.append((code, parse_stock_qty(qty_raw)))
        return parsed
    finally:
        workbook.close()


def apply_stock_rows(rows: list[tuple[str, int | None]]) -> StockImportReport:
    """Update ``SKU.stock_qty`` / ``stock_qty_ma`` for known codes; ignore unknown.

    Catalog SKUs absent from ``rows`` keep their previous quantities.
    Duplicate артикул in the file: last row wins (per bucket).

    A row whose article contains ``4-20 mA`` / ``4-20mA`` updates
    ``stock_qty_ma`` and does **not** overwrite ``stock_qty``. A base row
    for the same key sets ``stock_qty`` and zeros ``stock_qty_ma`` unless a
    mA row is also present.

    Matching: exact / casefold, plus BV body aliases
    (``BV232-A`` → ``8100-bv232a``). H81xx-BV* kit codes do not alias to
    bare BV bodies. Bare 1C kit articles (``H8121-BV265``) update **all**
    electrical editions (``…-24A`` … ``…-230DS``). Same for bare
    ``H8205-LAV280`` / ``H8205-LAV280ST`` → all matching electrical editions.

    Args:
        rows: Parsed ``(sku_code, qty_or_None)`` pairs.

    Returns:
        Import counters.
    """
    report = StockImportReport()
    if not rows:
        return report

    base_by_key: dict[str, tuple[str, int | None]] = {}
    ma_by_key: dict[str, tuple[str, int | None]] = {}
    for code, qty in rows:
        raw = code.strip()
        if not raw:
            report.skipped_blank += 1
            continue
        key = normalize_stock_article_key(raw)
        if not key:
            report.skipped_blank += 1
            continue
        if stock_article_is_ma_option(raw):
            ma_by_key[key] = (raw, qty)
        else:
            base_by_key[key] = (raw, qty)

    skus = list(
        SKU.objects.only(
            "id",
            "sku_code",
            "stock_qty",
            "stock_qty_ma",
            "stock_updated_at",
        ),
    )
    sku_by_key: dict[str, SKU] = {}
    skus_by_bare: dict[str, list[SKU]] = {}
    for sku in skus:
        key = normalize_stock_article_key(sku.sku_code)
        if not key:
            continue
        # Prefer first SKU for a key; identical aliases should be unique.
        sku_by_key.setdefault(key, sku)
        for bare_fn in (h81_kit_bare_stock_key, h8205_lav_bare_stock_key):
            bare = bare_fn(key)
            if bare is not None and bare != key:
                skus_by_bare.setdefault(bare, []).append(sku)

    now = timezone.now()
    to_update: list[SKU] = []
    seen_ids: set[int] = set()

    def _touch(sku: SKU) -> None:
        sku.stock_updated_at = now
        if sku.pk in seen_ids:
            return
        to_update.append(sku)
        seen_ids.add(sku.pk)

    for key in set(base_by_key) | set(ma_by_key):
        raw_code = (base_by_key.get(key) or ma_by_key[key])[0]
        targets = _resolve_stock_target_skus(key, sku_by_key, skus_by_bare)
        if not targets:
            report.ignored_unknown += 1
            if len(report.unknown_codes) < 20:
                report.unknown_codes.append(raw_code)
            continue
        base_entry = base_by_key.get(key)
        ma_entry = ma_by_key.get(key)
        if base_entry is not None and base_entry[1] is None:
            report.bad_rows += 1
        if ma_entry is not None and ma_entry[1] is None:
            report.bad_rows += 1
        applied = False
        if base_entry is not None and base_entry[1] is not None:
            for sku in targets:
                sku.stock_qty = base_entry[1]
                if ma_entry is None:
                    sku.stock_qty_ma = 0
                _touch(sku)
            applied = True
        if ma_entry is not None and ma_entry[1] is not None:
            for sku in targets:
                sku.stock_qty_ma = ma_entry[1]
                _touch(sku)
            applied = True
        if not applied:
            continue

    if to_update:
        with transaction.atomic():
            SKU.objects.bulk_update(
                to_update,
                ["stock_qty", "stock_qty_ma", "stock_updated_at"],
            )
        report.updated = len(to_update)
        logger.info(
            "stock_import updated=%s ignored_unknown=%s bad_rows=%s",
            report.updated,
            report.ignored_unknown,
            report.bad_rows,
        )
    return report


def _resolve_stock_target_skus(
    key: str,
    sku_by_key: dict[str, SKU],
    skus_by_bare: dict[str, list[SKU]],
) -> list[SKU]:
    """Map one stock key to one SKU or all H81/H8205 electrical editions."""
    for bare_fn in (h81_kit_bare_stock_key, h8205_lav_bare_stock_key):
        bare = bare_fn(key)
        if bare is not None and key == bare:
            editions = skus_by_bare.get(bare)
            if editions:
                return list(editions)
    sku = sku_by_key.get(key)
    return [sku] if sku is not None else []


def import_stock_xlsx(file_obj: BinaryIO | BytesIO) -> StockImportReport:
    """Parse an XLSX stream and apply stock quantities to catalog SKUs.

    Args:
        file_obj: Binary workbook (``.xlsx``).

    Returns:
        Import report.

    Raises:
        StockImportError: invalid workbook / headers.
    """
    return apply_stock_rows(parse_stock_rows(file_obj))


def build_stock_template_xlsx() -> bytes:
    """Return a minimal template workbook (Артикул | Свободно)."""
    book = Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Остатки"
    sheet.append(["Артикул", "Свободно"])
    sheet.append(["DA5FU24-D", 10])
    sheet.append(["BV232-A", 5])
    sheet.append(["HVD-3F24-ST", 0])
    buf = BytesIO()
    book.save(buf)
    return buf.getvalue()
